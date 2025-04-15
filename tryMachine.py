import websocket
import json
import threading
import os
from openpyxl import Workbook, load_workbook

# Variables globales para almacenar token, sesión, headset, workbook, worksheets y la instancia ws
cortex_token = None
session_id = None
headset_id = None
wb = None            # Workbook
ws_all = None        # Worksheet donde se guardan TODOS los comandos (incluyendo neutral)
ws_nonneutral = None # Worksheet donde se guardan los comandos que no sean "neutral"
ws_app = None        # Instancia global de WebSocketApp
filename = "MentalComandsSignals.xlsx"  # Nombre del archivo Excel

# Nombres de las hojas
sheet_all = "Comandos con Neutral"
sheet_non = "Comandos sin Neutral"

def crear_excel(filename=filename):
    """
    Abre el archivo Excel si existe; de lo contrario, crea uno nuevo con dos hojas:
    una para guardar todos los comandos (incluido "neutral") y otra para guardar solo los no neutral.
    """
    global wb, ws_all, ws_nonneutral
    if os.path.exists(filename):
        wb = load_workbook(filename)
        # Verificar o crear la hoja para todos los comandos
        if sheet_all in wb.sheetnames:
            ws_all = wb[sheet_all]
        else:
            ws_all = wb.create_sheet(sheet_all)
            ws_all.append(["Señal", "Potencia numérica de la señal", "Potencia categórica de la señal"])
        # Verificar o crear la hoja para comandos sin neutral
        if sheet_non in wb.sheetnames:
            ws_nonneutral = wb[sheet_non]
        else:
            ws_nonneutral = wb.create_sheet(sheet_non)
            ws_nonneutral.append(["Señal", "Potencia numérica de la señal", "Potencia categórica de la señal"])
        print("Archivo existente abierto:", filename)
    else:
        wb = Workbook()
        # Configurar la hoja de todos los comandos renombrando la hoja activa
        ws_all = wb.active
        ws_all.title = sheet_all
        ws_all.append(["Señal", "Potencia numérica de la señal", "Potencia categórica de la señal"])
        # Crear la hoja para comandos sin neutral
        ws_nonneutral = wb.create_sheet(sheet_non)
        ws_nonneutral.append(["Señal", "Potencia numérica de la señal", "Potencia categórica de la señal"])
        print("Nuevo archivo creado:", filename)
    return wb, ws_all, ws_nonneutral

def guardar_excel(wb, nombre_archivo=filename):
    wb.save(nombre_archivo)
    print("Archivo Excel guardado:", nombre_archivo)

def on_message(ws, message):
    global cortex_token, session_id, headset_id, ws_all, ws_nonneutral
    data = json.loads(message)
    
    # 1. Respuesta a la autorización
    if data.get("id") == 1 and "result" in data:
        cortex_token = data["result"]["cortexToken"]
        print("Cortex Token obtenido:", cortex_token)
        
        # Enviar solicitud queryHeadsets
        query_headset_request = {
            "jsonrpc": "2.0",
            "method": "queryHeadsets",
            "params": {},
            "id": 2
        }
        ws.send(json.dumps(query_headset_request))
    
    # 2. Respuesta a queryHeadsets: obtener el headset id
    elif data.get("id") == 2 and "result" in data:
        if data["result"]:
            headset_id = data["result"][0]["id"]
            print("Headset encontrado:", headset_id)
            
            # Crear sesión con el headset
            create_session_request = {
                "jsonrpc": "2.0",
                "method": "createSession",
                "params": {
                    "cortexToken": cortex_token,
                    "headset": headset_id,
                    "status": "active"
                },
                "id": 3
            }
            ws.send(json.dumps(create_session_request))
        else:
            print("No se encontraron headsets disponibles.")
    
    # 3. Respuesta a createSession
    elif data.get("id") == 3 and "result" in data:
        session_id = data["result"]["id"]
        print("Sesión creada con ID:", session_id)
        
        # Suscribirse al stream de comandos mentales ("com")
        subscribe_request = {
            "jsonrpc": "2.0",
            "method": "subscribe",
            "params": {
                "cortexToken": cortex_token,
                "session": session_id,
                "streams": ["com"]
            },
            "id": 4
        }
        ws.send(json.dumps(subscribe_request))
    
    # 4. Respuesta a la suscripción
    elif data.get("id") == 4 and "result" in data:
        print("Suscripción a comandos mentales completada.")
    
    # 5. Procesamiento de datos de comandos mentales
    if "com" in data:
        # Suponiendo que data["com"] es una lista: [accion, potencia]
        accion = data["com"][0]
        potencia = data["com"][1]
        
        # Categorizar la potencia (ejemplo: thresholds en 0.30 y 0.70)
        categoria = "Bajo"
        if potencia > 0.30:
            categoria = "Medio"
        if potencia > 0.70:
            categoria = "Alto"
        
        # Mostrar la potencia según el comando
        if accion == "neutral":
            print("<-> Potencia: ", potencia * 100, '%')
        elif accion == "left":
            print("<-- Potencia: ", potencia * 100, '%')
        elif accion == "right":
            print("--> Potencia: ", potencia * 100, '%')
        
        # Guardar la fila en la hoja que guarda TODOS los comandos (incluido "neutral")
        ws_all.append([accion, potencia, categoria])
        # Guardar en la hoja que NO guarda "neutral" solo si el comando es distinto de "neutral"
        if accion != 'neutral':
            ws_nonneutral.append([accion, potencia, categoria])

def on_error(ws, error):
    print("Error:", error)

def on_close(ws, close_status_code, close_msg):
    guardar_excel(wb)
    print("Conexión cerrada")

def on_open(ws):
    # Enviar solicitud de autorización
    auth_request = {
        "jsonrpc": "2.0",
        "method": "authorize",
        "params": {
            "clientId": "RxcaORHl6m1MYTi3tYB1bBA4sWxjqdcMDynGuWlL",
            "clientSecret": "v1AmcuvPeDRGwNYHhC3R5ycWf9MSTs6hvy4Na6RrJIXFouwLQCWqyRua7lKrEjn3C6a10g6yGoRXBEGgf5SzuXUGgBZTgkvu4J8qYhzQKZZczw5cTPdsHapGV9CZWFGJ",
            "debit": 0
        },
        "id": 1
    }
    ws.send(json.dumps(auth_request))

def run_ws():
    global ws_app
    ws_url = "wss://127.0.0.1:6868"  # Verifica que este endpoint sea correcto
    ws_app = websocket.WebSocketApp(ws_url,
                                    on_open=on_open,
                                    on_message=on_message,
                                    on_error=on_error,
                                    on_close=on_close)
    ws_app.run_forever()

if __name__ == "__main__":
    # Abrir el archivo Excel existente o crearlo si no existe y obtener ambas hojas
    wb, ws_all, ws_nonneutral = crear_excel()

    # Iniciar el WebSocket en un hilo separado
    ws_thread = threading.Thread(target=run_ws)
    ws_thread.start()

    # Esperar a que el usuario presione Enter para cerrar la conexión
    input("Press Enter to close the connection...\n")

    print("Closing connection... (please wait for the on_close callback)")
    
    # Cerrar la conexión llamando al método close() del objeto ws_app
    if ws_app:
        ws_app.close()

    # Esperar a que el hilo termine
    ws_thread.join()
